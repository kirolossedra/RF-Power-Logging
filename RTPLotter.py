#!/usr/bin/env python3

import os
import time
import struct
import threading
from datetime import datetime

import numpy as np
import serial
from serial.tools import list_ports
from openpyxl import Workbook

import tkinter as tk
from tkinter import ttk, messagebox

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.backends.backend_pdf import PdfPages


VID = 0x0483
PID = 0x5740


def getport():
    for device in list_ports.comports():
        if device.vid == VID and device.pid == PID:
            return device.device
    raise OSError("tinySA not found")


def format_dt(dt_obj):
    return dt_obj.strftime("%Y-%m-%d %H:%M:%S")


def sanitize_sheet_name(name):
    invalid = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in invalid:
        name = name.replace(ch, "_")
    return name[:31]


class TinySA:
    def __init__(self, port):
        self.ser = serial.Serial(port, 115200, timeout=1)
        self.lock = threading.Lock()

    def close(self):
        with self.lock:
            try:
                if self.ser and self.ser.is_open:
                    self.ser.close()
            except Exception:
                pass

    def _read_prompt(self):
        data = self.ser.read_until(b"ch> ")
        if not data.endswith(b"ch> "):
            raise RuntimeError("tinySA prompt not received")

    def _flush_buffers(self):
        try:
            self.ser.reset_input_buffer()
            self.ser.reset_output_buffer()
        except Exception:
            pass

        end_time = time.time() + 0.3
        while time.time() < end_time:
            try:
                waiting = self.ser.in_waiting
            except Exception:
                waiting = 0

            if waiting > 0:
                try:
                    self.ser.read(waiting)
                except Exception:
                    break
                time.sleep(0.02)
            else:
                break

    def scan(self, f_low, f_high, points, rbw, stop_event=None):
        with self.lock:
            if not self.ser or not self.ser.is_open:
                raise RuntimeError("Serial port is closed")

            if stop_event is not None and stop_event.is_set():
                raise RuntimeError("Scan stopped")

            self._flush_buffers()

            if rbw == 0:
                rbw_k = (f_high - f_low) * 7e-6
            else:
                rbw_k = rbw / 1e3

            rbw_k = max(3, min(600, rbw_k))

            self.ser.write(f"rbw {int(rbw_k)}\r".encode())
            self._read_prompt()

            timeout = ((f_high - f_low) / 20e3) / (rbw_k ** 2) + points / 500 + 1
            self.ser.timeout = max(2, timeout * 2)

            if stop_event is not None and stop_event.is_set():
                raise RuntimeError("Scan stopped")

            self.ser.write(f"scanraw {int(f_low)} {int(f_high)} {int(points)}\r".encode())

            head = self.ser.read_until(b"{")
            if not head.endswith(b"{"):
                raise RuntimeError("tinySA did not start scan payload")

            raw = self.ser.read_until(b"}ch> ")
            if not raw.endswith(b"}ch> "):
                raise RuntimeError("tinySA returned incomplete scan payload")

            payload = raw[:-5]

            expected_len = points * 3
            if len(payload) != expected_len:
                raise RuntimeError(
                    f"Unexpected payload size: got {len(payload)}, expected {expected_len}"
                )

            try:
                data = struct.unpack("<" + "xH" * points, payload)
            except struct.error as e:
                raise RuntimeError(f"Failed to decode scan payload: {e}")

            arr = np.array(data, dtype=np.uint16)

            scale = 128
            power = arr / 32 - scale

            try:
                self.ser.timeout = 1
                self.ser.write(b"rbw auto\r")
                self._read_prompt()
            except Exception:
                pass

            return power


class RTPlotter:
    def __init__(self, root):
        self.root = root
        self.root.title("tinySA RT Plotter")

        self.running = False
        self.worker = None
        self.stop_event = threading.Event()
        self.tiny = None

        self.data_lock = threading.Lock()
        self.freq = None
        self.power = None
        self.scan_count = 0

        self.f_low = None
        self.f_high = None
        self.num_points = None
        self.rbw_value = None

        self.auto_run_folder = None
        self.auto_run_started_at = None
        self.last_sweep_duration = None
        self.completed_sweep_durations = []

        self.auto_excel_path = None
        self.auto_pdf_path = None
        self.auto_records = []

        self.build_gui()
        self.build_plot()

    def build_gui(self):
        frame = ttk.Frame(self.root, padding=10)
        frame.pack(fill=tk.X)

        ttk.Label(frame, text="Start Hz").grid(row=0, column=0, padx=3, pady=3)
        self.start = tk.StringVar(value="2300000000")
        ttk.Entry(frame, textvariable=self.start, width=15).grid(row=0, column=1, padx=3, pady=3)

        ttk.Label(frame, text="End Hz").grid(row=0, column=2, padx=3, pady=3)
        self.end = tk.StringVar(value="2500000000")
        ttk.Entry(frame, textvariable=self.end, width=15).grid(row=0, column=3, padx=3, pady=3)

        ttk.Label(frame, text="Points").grid(row=0, column=4, padx=3, pady=3)
        self.points_var = tk.StringVar(value="401")
        ttk.Entry(frame, textvariable=self.points_var, width=8).grid(row=0, column=5, padx=3, pady=3)

        ttk.Label(frame, text="RBW").grid(row=0, column=6, padx=3, pady=3)
        self.rbw = tk.StringVar(value="0")
        ttk.Entry(frame, textvariable=self.rbw, width=8).grid(row=0, column=7, padx=3, pady=3)

        ttk.Label(frame, text="Run Mode").grid(row=1, column=0, padx=3, pady=3)
        self.run_mode = tk.StringVar(value="Continuous Live")
        ttk.Combobox(
            frame,
            textvariable=self.run_mode,
            values=["Continuous Live", "Automatic Sweep Mode"],
            state="readonly",
            width=20
        ).grid(row=1, column=1, padx=3, pady=3)

        ttk.Label(frame, text="Sweeps").grid(row=1, column=2, padx=3, pady=3)
        self.auto_sweeps = tk.StringVar(value="5")
        ttk.Entry(frame, textvariable=self.auto_sweeps, width=8).grid(row=1, column=3, padx=3, pady=3)

        ttk.Label(frame, text="Wait(s)").grid(row=1, column=4, padx=3, pady=3)
        self.wait_between_sweeps = tk.StringVar(value="2")
        ttk.Entry(frame, textvariable=self.wait_between_sweeps, width=8).grid(row=1, column=5, padx=3, pady=3)

        self.port = tk.StringVar()

        ttk.Button(frame, text="Auto Detect", command=self.detect).grid(row=2, column=0, padx=3, pady=3)

        self.start_btn = ttk.Button(frame, text="Start", command=self.start_scan)
        self.start_btn.grid(row=2, column=1, padx=3, pady=3)

        self.stop_btn = ttk.Button(frame, text="Stop", command=self.stop_scan, state=tk.DISABLED)
        self.stop_btn.grid(row=2, column=2, padx=3, pady=3)

        ttk.Button(frame, text="Snapshot", command=self.take_snapshot).grid(row=2, column=3, padx=3, pady=3)
        ttk.Button(frame, text="Estimate Auto Time", command=self.show_estimate).grid(row=2, column=4, padx=3, pady=3)

        self.status = tk.StringVar(value="Idle")
        ttk.Label(self.root, textvariable=self.status, relief=tk.SUNKEN, anchor="w").pack(fill=tk.X)

    def build_plot(self):
        self.fig = Figure(figsize=(10, 5))
        self.ax = self.fig.add_subplot(111)

        self.raw_line, = self.ax.plot([], [], label="Sweep")

        self.ax.set_xlabel("Frequency (Hz)")
        self.ax.set_ylabel("dBm")
        self.ax.grid(True)
        self.ax.legend()

        self.canvas = FigureCanvasTkAgg(self.fig, master=self.root)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def detect(self):
        try:
            p = getport()
            self.port.set(p)
            self.status.set(f"tinySA detected on {p}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def clear_plot_data(self):
        with self.data_lock:
            self.freq = None
            self.power = None
            self.scan_count = 0
            self.last_sweep_duration = None
            self.completed_sweep_durations = []
            self.auto_records = []
            self.auto_excel_path = None
            self.auto_pdf_path = None

        self.raw_line.set_data([], [])
        self.canvas.draw_idle()

    def parse_inputs(self):
        port = self.port.get().strip() or getport()
        f_low = float(self.start.get())
        f_high = float(self.end.get())
        num_points = int(self.points_var.get())
        rbw_value = float(self.rbw.get())

        if f_low >= f_high:
            raise ValueError("Start frequency must be less than end frequency")
        if num_points < 2:
            raise ValueError("Points must be at least 2")
        if rbw_value < 0:
            raise ValueError("RBW must be 0 or positive")

        return port, f_low, f_high, num_points, rbw_value

    def parse_auto_inputs(self):
        total_sweeps = int(self.auto_sweeps.get())
        wait_seconds = float(self.wait_between_sweeps.get())

        if total_sweeps < 1:
            raise ValueError("Sweeps must be at least 1")
        if wait_seconds < 0:
            raise ValueError("Wait time must be 0 or positive")

        return total_sweeps, wait_seconds

    def start_scan(self):
        if self.running:
            self.stop_scan()

        try:
            port, self.f_low, self.f_high, self.num_points, self.rbw_value = self.parse_inputs()
            total_sweeps, wait_seconds = self.parse_auto_inputs()
        except Exception as e:
            messagebox.showerror("Input Error", str(e))
            return

        try:
            self.tiny = TinySA(port)
        except Exception as e:
            messagebox.showerror("Serial Error", str(e))
            return

        self.clear_plot_data()

        with self.data_lock:
            self.freq = np.linspace(self.f_low, self.f_high, self.num_points)

        self.stop_event.clear()
        self.running = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        run_mode = self.run_mode.get()

        if run_mode == "Automatic Sweep Mode":
            self.status.set("Automatic sweep mode running...")
            self.worker = threading.Thread(
                target=self.auto_sweep_loop,
                args=(total_sweeps, wait_seconds),
                daemon=True
            )
        else:
            self.status.set("Scanning...")
            self.worker = threading.Thread(target=self.live_loop, daemon=True)

        self.worker.start()

    def stop_scan(self):
        self.running = False
        self.stop_event.set()

        tiny = self.tiny
        self.tiny = None

        if tiny is not None:
            try:
                tiny.close()
            except Exception:
                pass

        worker = self.worker
        if worker and worker.is_alive() and worker is not threading.current_thread():
            worker.join(timeout=2)

        self.worker = None

        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status.set("Stopped")

    def update_data_after_sweep(self, power, duration_seconds):
        with self.data_lock:
            self.power = power.copy()
            self.scan_count += 1
            self.last_sweep_duration = duration_seconds
            self.completed_sweep_durations.append(duration_seconds)
            count = self.scan_count

        return count

    def live_loop(self):
        while not self.stop_event.is_set():
            try:
                tiny = self.tiny
                if tiny is None:
                    break

                t0 = time.time()
                power = tiny.scan(
                    self.f_low,
                    self.f_high,
                    self.num_points,
                    self.rbw_value,
                    stop_event=self.stop_event
                )

                if self.stop_event.is_set():
                    break

                dt = time.time() - t0
                count = self.update_data_after_sweep(power, dt)

                self.root.after(
                    0,
                    lambda c=count, d=dt: self.status.set(
                        f"live sweep {c} finished in {d:.2f}s"
                    )
                )

            except Exception as e:
                if self.stop_event.is_set():
                    break

                err = str(e)

                def handle_error():
                    self.status.set(f"Error: {err}")
                    self.stop_scan()
                    messagebox.showerror("Scan Error", err)

                self.root.after(0, handle_error)
                break

    def create_run_folder(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = os.path.join(script_dir, timestamp)
        os.makedirs(folder, exist_ok=True)
        self.auto_run_folder = folder
        self.auto_run_started_at = timestamp
        self.auto_excel_path = os.path.join(folder, f"{timestamp}_sweeps.xlsx")
        self.auto_pdf_path = os.path.join(folder, f"{timestamp}_sweeps_report.pdf")
        return folder, timestamp

    def save_sweep_image(self, folder, sweep_index, freq, power, sweep_start_dt, sweep_end_dt):
        image_name = f"sweep_{sweep_index:03d}_{sweep_start_dt.strftime('%Y%m%d_%H%M%S')}.png"
        image_path = os.path.join(folder, image_name)

        fig = Figure(figsize=(10, 5))
        ax = fig.add_subplot(111)

        ax.plot(freq, power, label="Sweep")
        ax.set_xlabel("Frequency (Hz)")
        ax.set_ylabel("dBm")
        ax.set_title(
            f"Sweep {sweep_index} | Start: {format_dt(sweep_start_dt)} | End: {format_dt(sweep_end_dt)}"
        )
        ax.grid(True)
        ax.legend()

        fig.tight_layout()
        fig.savefig(image_path, dpi=150)
        return image_path

    def build_workbook(self):
        workbook = Workbook()

        summary = workbook.active
        summary.title = "Summary"
        summary["A1"] = "Sweep_Index"
        summary["B1"] = "Start_Time"
        summary["C1"] = "End_Time"
        summary["D1"] = "Duration_s"
        summary["E1"] = "Image_File"

        return workbook

    def add_sweep_sheet(self, workbook, sweep_index, freq, power, sweep_start_dt, sweep_end_dt, duration_seconds, wait_seconds):
        sheet_name = sanitize_sheet_name(f"Sweep_{sweep_index}")
        sheet = workbook.create_sheet(title=sheet_name)

        sheet["A1"] = "Sweep_Index"
        sheet["B1"] = sweep_index

        sheet["A2"] = "Start_Time"
        sheet["B2"] = format_dt(sweep_start_dt)

        sheet["A3"] = "End_Time"
        sheet["B3"] = format_dt(sweep_end_dt)

        sheet["A4"] = "Sweep_Duration_s"
        sheet["B4"] = float(duration_seconds)

        sheet["A5"] = "Wait_Between_Sweeps_s"
        sheet["B5"] = float(wait_seconds)

        sheet["A7"] = "Frequency_Hz"
        sheet["B7"] = "Power_dBm"

        for idx, f in enumerate(freq, start=8):
            sheet.cell(row=idx, column=1, value=float(f))
            sheet.cell(row=idx, column=2, value=float(power[idx - 8]))

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 24

    def update_summary_sheet(self, workbook, records):
        if "Summary" not in workbook.sheetnames:
            return

        sheet = workbook["Summary"]

        max_existing_rows = sheet.max_row
        if max_existing_rows > 1:
            sheet.delete_rows(2, max_existing_rows - 1)

        for row_idx, record in enumerate(records, start=2):
            sheet.cell(row=row_idx, column=1, value=int(record["index"]))
            sheet.cell(row=row_idx, column=2, value=record["start"])
            sheet.cell(row=row_idx, column=3, value=record["end"])
            sheet.cell(row=row_idx, column=4, value=float(record["duration"]))
            sheet.cell(row=row_idx, column=5, value=os.path.basename(record["image_path"]))

        for col in ["A", "B", "C", "D", "E"]:
            sheet.column_dimensions[col].width = 24

    def save_summary_pdf(self, pdf_path, records, run_started_at, run_finished_at, freq):
        with PdfPages(pdf_path) as pdf:
            summary_fig = Figure(figsize=(11.69, 8.27))
            ax = summary_fig.add_subplot(111)
            ax.axis("off")

            title_lines = [
                "Automatic Sweep Report",
                f"Run folder: {os.path.basename(os.path.dirname(pdf_path))}",
                f"Run started: {format_dt(run_started_at)}",
                f"Run finished: {format_dt(run_finished_at)}",
                f"Frequency range: {freq[0]:.0f} Hz to {freq[-1]:.0f} Hz",
                f"Total sweeps saved: {len(records)}",
            ]

            y = 0.95
            for line in title_lines:
                ax.text(0.03, y, line, fontsize=12, va="top")
                y -= 0.055

            y -= 0.02
            ax.text(0.03, y, "Sweep Summary", fontsize=12, va="top")
            y -= 0.05

            header = f"{'Sweep':<8}{'Start Time':<24}{'End Time':<24}{'Duration (s)':<14}"
            ax.text(0.03, y, header, family="monospace", fontsize=10, va="top")
            y -= 0.035

            for record in records:
                line = (
                    f"{record['index']:<8}"
                    f"{record['start']:<24}"
                    f"{record['end']:<24}"
                    f"{record['duration']:<14.2f}"
                )
                ax.text(0.03, y, line, family="monospace", fontsize=9, va="top")
                y -= 0.03

                if y < 0.06:
                    pdf.savefig(summary_fig)
                    summary_fig = Figure(figsize=(11.69, 8.27))
                    ax = summary_fig.add_subplot(111)
                    ax.axis("off")
                    y = 0.95

            summary_fig.tight_layout()
            pdf.savefig(summary_fig)

            for record in records:
                fig = Figure(figsize=(11.69, 8.27))
                ax = fig.add_subplot(111)

                ax.plot(record["freq"], record["power"], label="Sweep")
                ax.set_xlabel("Frequency (Hz)")
                ax.set_ylabel("dBm")
                ax.set_title(
                    f"Sweep {record['index']} | Start: {record['start']} | End: {record['end']}"
                )
                ax.grid(True)
                ax.legend()
                fig.tight_layout()
                pdf.savefig(fig)

    def save_incremental_outputs(self, workbook, records, run_started_dt, run_finished_dt):
        if not records or self.auto_excel_path is None or self.auto_pdf_path is None:
            return

        self.update_summary_sheet(workbook, records)
        workbook.save(self.auto_excel_path)
        self.save_summary_pdf(
            pdf_path=self.auto_pdf_path,
            records=records,
            run_started_at=run_started_dt,
            run_finished_at=run_finished_dt,
            freq=records[0]["freq"]
        )

    def auto_sweep_loop(self, total_sweeps, wait_seconds):
        workbook = None
        records = []
        folder = None
        run_started_dt = datetime.now()

        try:
            tiny = self.tiny
            if tiny is None:
                raise RuntimeError("tinySA not initialized")

            folder, run_stamp = self.create_run_folder()

            workbook = self.build_workbook()

            for sweep_idx in range(1, total_sweeps + 1):
                if self.stop_event.is_set():
                    break

                sweep_start_dt = datetime.now()
                t0 = time.time()

                power = tiny.scan(
                    self.f_low,
                    self.f_high,
                    self.num_points,
                    self.rbw_value,
                    stop_event=self.stop_event
                )

                if self.stop_event.is_set():
                    break

                duration_seconds = time.time() - t0
                sweep_end_dt = datetime.now()

                count = self.update_data_after_sweep(power, duration_seconds)

                with self.data_lock:
                    freq = None if self.freq is None else self.freq.copy()

                if freq is None:
                    raise RuntimeError("Frequency axis is not initialized")

                image_path = self.save_sweep_image(
                    folder=folder,
                    sweep_index=sweep_idx,
                    freq=freq,
                    power=power,
                    sweep_start_dt=sweep_start_dt,
                    sweep_end_dt=sweep_end_dt
                )

                self.add_sweep_sheet(
                    workbook=workbook,
                    sweep_index=sweep_idx,
                    freq=freq,
                    power=power,
                    sweep_start_dt=sweep_start_dt,
                    sweep_end_dt=sweep_end_dt,
                    duration_seconds=duration_seconds,
                    wait_seconds=wait_seconds
                )

                avg_duration = sum(self.completed_sweep_durations) / len(self.completed_sweep_durations)
                remaining_sweeps = total_sweeps - sweep_idx
                estimated_remaining_seconds = remaining_sweeps * avg_duration + remaining_sweeps * wait_seconds

                records.append({
                    "index": sweep_idx,
                    "start": format_dt(sweep_start_dt),
                    "end": format_dt(sweep_end_dt),
                    "duration": duration_seconds,
                    "freq": freq.copy(),
                    "power": power.copy(),
                    "image_path": image_path,
                })

                with self.data_lock:
                    self.auto_records = list(records)

                self.save_incremental_outputs(
                    workbook=workbook,
                    records=records,
                    run_started_dt=run_started_dt,
                    run_finished_dt=sweep_end_dt
                )

                self.root.after(
                    0,
                    lambda i=sweep_idx, n=total_sweeps, d=duration_seconds,
                           avg=avg_duration, rem=estimated_remaining_seconds, c=count: self.status.set(
                        f"auto sweep {i}/{n} finished in {d:.2f}s | avg {avg:.2f}s | remaining ~{rem:.2f}s | scan {c}"
                    )
                )

                if sweep_idx < total_sweeps:
                    wait_start = time.time()
                    while True:
                        if self.stop_event.is_set():
                            break

                        elapsed = time.time() - wait_start
                        remaining = wait_seconds - elapsed
                        if remaining <= 0:
                            break

                        time.sleep(min(0.1, remaining))

                    if self.stop_event.is_set():
                        break

            run_finished_dt = datetime.now()

            if workbook is not None and records:
                self.save_incremental_outputs(
                    workbook=workbook,
                    records=records,
                    run_started_dt=run_started_dt,
                    run_finished_dt=run_finished_dt
                )

                self.root.after(
                    0,
                    lambda: self.status.set(
                        f"Automatic sweep mode finished | folder: {folder} | PDF: {os.path.basename(self.auto_pdf_path)} | Excel: {os.path.basename(self.auto_excel_path)}"
                    )
                )
            elif folder is not None:
                self.root.after(
                    0,
                    lambda: self.status.set(
                        f"Automatic sweep mode stopped before first saved sweep | folder: {folder}"
                    )
                )

        except Exception as e:
            if self.stop_event.is_set():
                return

            err = str(e)

            def handle_error():
                self.status.set(f"Error: {err}")
                self.stop_scan()
                messagebox.showerror("Automatic Sweep Error", err)

            self.root.after(0, handle_error)
            return

        finally:
            self.root.after(0, self.finish_worker_state)

    def finish_worker_state(self):
        self.running = False

        tiny = self.tiny
        self.tiny = None
        if tiny is not None:
            try:
                tiny.close()
            except Exception:
                pass

        self.worker = None
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

    def show_estimate(self):
        try:
            total_sweeps, wait_seconds = self.parse_auto_inputs()

            with self.data_lock:
                durations = self.completed_sweep_durations.copy()

            if durations:
                avg_duration = sum(durations) / len(durations)
                total_estimated = total_sweeps * avg_duration + max(total_sweeps - 1, 0) * wait_seconds
                msg = (
                    f"Estimated sweep time based on measured average:\n\n"
                    f"Average per sweep: {avg_duration:.2f} s\n"
                    f"Total sweeps: {total_sweeps}\n"
                    f"Wait between sweeps: {wait_seconds:.2f} s\n"
                    f"Estimated total automatic run: {total_estimated:.2f} s"
                )
            else:
                msg = (
                    "No completed sweep yet, so exact sweep duration is not known.\n\n"
                    "Run one sweep first, then this button will estimate:\n"
                    "- time per sweep\n"
                    "- total automatic sweep time"
                )

            messagebox.showinfo("Automatic Sweep Estimate", msg)

        except Exception as e:
            messagebox.showerror("Estimate Error", str(e))

    def update_plot(self):
        with self.data_lock:
            freq = None if self.freq is None else self.freq.copy()
            power = None if self.power is None else self.power.copy()

        if freq is not None and power is not None:
            self.raw_line.set_data(freq, power)
            self.raw_line.set_visible(True)

            self.ax.set_xlim(freq[0], freq[-1])

            ymin = float(np.min(power)) - 5
            ymax = float(np.max(power)) + 5

            if ymin == ymax:
                ymin -= 1
                ymax += 1

            self.ax.set_ylim(ymin, ymax)
            self.canvas.draw_idle()

        self.root.after(100, self.update_plot)

    def save_single_snapshot_excel(self, filepath_base, freq, power):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SweepData"

        sheet["A1"] = "Frequency_Hz"
        sheet["B1"] = "Power_dBm"
        sheet["C1"] = "Export_Timestamp"
        sheet["D1"] = "Scan_Count"

        export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for idx, f in enumerate(freq, start=2):
            sheet.cell(row=idx, column=1, value=float(f))
            sheet.cell(row=idx, column=2, value=float(power[idx - 2]))
            sheet.cell(row=idx, column=3, value=export_time)
            sheet.cell(row=idx, column=4, value=int(self.scan_count))

        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 22
        sheet.column_dimensions["D"].width = 12

        excel_path = filepath_base + ".xlsx"
        workbook.save(excel_path)
        return excel_path

    def take_snapshot(self):
        try:
            with self.data_lock:
                freq = None if self.freq is None else self.freq.copy()
                power = None if self.power is None else self.power.copy()
                scan_count = self.scan_count

            if freq is None or power is None:
                raise RuntimeError("No sweep data available yet")

            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_name = f"snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            image_path = os.path.join(script_dir, base_name + ".png")

            fig = Figure(figsize=(10, 5))
            ax = fig.add_subplot(111)
            ax.plot(freq, power, label="Sweep")
            ax.set_xlabel("Frequency (Hz)")
            ax.set_ylabel("dBm")
            ax.set_title(f"Snapshot | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ax.grid(True)
            ax.legend()
            fig.tight_layout()
            fig.savefig(image_path, dpi=150)

            excel_path = self.save_single_snapshot_excel(
                os.path.join(script_dir, base_name),
                freq,
                power
            )

            self.status.set(
                f"Snapshot saved: {image_path} | Excel saved: {excel_path} | scan {scan_count}"
            )

        except Exception as e:
            messagebox.showerror("Snapshot Error", str(e))

    def on_close(self):
        self.stop_scan()
        self.root.destroy()


def main():
    root = tk.Tk()
    app = RTPlotter(root)
    root.after(100, app.update_plot)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
